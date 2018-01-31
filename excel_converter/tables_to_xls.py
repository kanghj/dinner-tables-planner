import sys
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import string
import itertools
import io


def make_workbook(persons, communities):
    wb = Workbook()
    ws = wb.active

    alphabet_first = [''] + [letter for letter in string.ascii_uppercase]
    column_names = [''.join(name)
                    for name in itertools.combinations_with_replacement(alphabet_first, 2)
                    ][1:]

    starting_col = 2
    for i, (table_name, members) in enumerate(communities.items()):
        column_name = column_names[i]
        ws[column_name + '1'] = 'Table ' + str(table_name).lstrip('=')
        for j, member in enumerate(members):
            ws[column_name + str(j + starting_col)] = persons[int(member) - 1].lstrip('=')

    return io.BytesIO(save_virtual_workbook(wb))




